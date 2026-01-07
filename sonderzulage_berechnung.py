import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import unicodedata

# Deutsche Monatsnamen
GERMAN_MONTHS = [
    "Dummy", "Januar", "Februar", "März", "April", "Mai", "Juni",
    "Juli", "August", "September", "Oktober", "November", "Dezember"
]

def get_german_month_name(month_number: int) -> str:
    return GERMAN_MONTHS[month_number]

# -------------------------------
# Fahrzeugart (ERWEITERT)
# -------------------------------
def define_art(value: int) -> str:
    if value in [602, 156]:
        return "Gigaliner"
    elif value in [350, 620]:
        return "Tandem"
    # Gliederzug inkl. neue LKW
    elif value in [520, 266, 458, 548, 541, 542, 543, 558]:
        return "Gliederzug"
    return "Unbekannt"

# -------------------------------
# Personalnummer-Zuordnung
# -------------------------------
name_to_personalnummer = {
    "Adler": {"Philipp": "00041450"},
    "Auer": {"Frank": "00020795"},
    "Batkowski": {"Tilo": "00046601"},
    "Benabbes": {"Badr": "00048980"},
    "Biebow": {"Thomas": "00042004"},
    "Bläsing": {"Elmar": "00049093"},
    "Bursian": {"Ronny": "00025714"},
    "Buth": {"Sven": "00046673"},
    "Böhnke": {"Marcel": "00020833"},
    "Carstensen": {"Martin": "00042412"},
    "Chege": {"Moses Gichuru": "00046106"},
    "Dammasch": {"Bernd": "00019297"},
    "Demuth": {"Harry": "00020796"},
    "Doroszkiewicz": {"Bogumil": "00049132"},
    "Dürr": {"Holger": "00039164"},
    "Effenberger": {"Sven": "00030807"},
    "Engel": {"Raymond": "00033429"},
    "Fechner": {"Danny": "00043696", "Klaus": "00038278"},
    "Findeklee": {"Bernd": "00020804"},
    "Flint": {"Henryk": "00042414"},
    "Fuhlbrügge": {"Justin": "00046289"},
    "Gehrmann": {"Rayk": "00046702"},
    "Gheonea": {"Costel-Daniel": "00054489"},
    "Glanz": {"Björn": "00041914"},
    "Gnech": {"Torsten": "00018613"},
    "Greve": {"Nicole": "00040760"},
    "Guthmann": {"Fred": "00018328"},
    "Hagen": {"Andy": "00020271"},
    "Hartig": {"Sebastian": "00044120"},
    "Haus": {"David": "00046101"},
    "Heeser": {"Bernd": "00041916"},
    "Helm": {"Philipp": "00046685"},
    "Henkel": {"Bastian": "00048187"},
    "Holtz": {"Torsten": "00021159"},
    "Hirdina": {"Christopher": "00053400"},
    "Hintz": {"Leif": "00054808"},
    "Hübner": {"Christian": "00054531"},
    "Janikiewicz": {"Radoslaw": "00042159"},
    "Kelling": {"Jonas Ole": "00044140"},
    "Kleiber": {"Lutz": "00026255"},
    "Klemkow": {"Ralf": "00040634"},
    "Kollmann": {"Steffen": "00040988"},
    "König": {"Heiko": "00036341"},
    "Krazewski": {"Cezary": "00039463"},
    "Krieger": {"Christian": "00049092"},
    "Krull": {"Benjamin": "00044192"},
    "Lange": {"Michael": "00035407"},
    "Lewandowski": {"Kamil": "00041044"},
    "Likoonski": {"Vladimir": "00044766"},
    "Linke": {"Erich": "00048377"},
    "Lefkih": {"Houssni": "00052293"},
    "Ludolf": {"Michel": "00048814"},
    "Lämmel": {"Patrick": "00052946"},
    "Marouni": {"Ayyoub": "00048986"},
    "Mintel": {"Mario": "00046686"},
    "Ohlenroth": {"Nadja": "00042114"},
    "Ohms": {"Torsten": "00019300"},
    "Okoth": {"Tedy Omondi": "00046107"},
    "Oszmian": {"Jacub": "00039464"},
    "Paul": {"Toralf": "00010490"},
    "Pabst": {"Torsten": "00021976"},
    "Pawlak": {"Bartosz": "00036381"},
    "Piepke": {"Torsten": "00021390"},
    "Plinke": {"Killian": "00044137"},
    "Pogodski": {"Enrico": "00046668"},
    "Postu": {"Mihai": "00051391"},
    "Quint": {"Stefan": "00035718"},
    "Rimba": {"Rimba Gona": "00046108"},
    "Rheinschmitt": {"Ronald": "00053356"},
    "Rudert": {"Kevin": "00052858"},
    "Rudolph": {"Yves": "00052855"},
    "Ruge": {"Fabian": "00054705"},
    "Sarwatka": {"Heiko": "00028747"},
    "Swietoslawski": {"Jacek": "00052955"},
    "Seredynski": {"Ireneusz": "00053452"},
    "Scheil": {"Eric-Rene": "00038579", "Rene": "00020851"},
    "Schlichting": {"Michael": "00021452"},
    "Schlutt": {"Hubert": "00020880", "Rene": "00042932"},
    "Schmieder": {"Steffen": "00046286"},
    "Schneider": {"Matthias": "00045495"},
    "Schulz": {"Julian": "00049130", "Stephan": "00041558"},
    "Singh": {"Jagtah": "00040902"},
    "Stoltz": {"Thorben": "00040991"},
    "Thal": {"Jannic": "00046006"},
    "Tumanow": {"Vasilli": "00045019"},
    "Wachnowski": {"Klaus": "00026019"},
    "Wendel": {"Danilo": "00048994"},
    "Waschitschek": {"Detlef": "00020436"},
    "Wille": {"Rene": "00021393"},
    "Wisniewski": {"Krzysztof": "00046550"},
    "Zander": {"Jan": "00042454"},
    "Zosel": {"Ingo": "00026303"},
}

# --- Namens-Normalisierung & robuster Lookup (VERBESSERT) ---
def _norm_simple(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("\u00a0", " ")                 # NBSP -> space
    s = unicodedata.normalize("NFKC", s)         # Unicode vereinheitlichen
    s = s.replace("-", " ")                      # Bindestrich -> Space
    s = (s.replace("ä", "ae")
           .replace("ö", "oe")
           .replace("ü", "ue")
           .replace("ß", "ss"))
    s = " ".join(s.split())
    return s

def get_personalnummer(nachname: str, vorname: str) -> str:
    n_key = _norm_simple(nachname)
    v_key = _norm_simple(vorname)

    hit = None
    for ln, inner in name_to_personalnummer.items():
        if _norm_simple(ln) == n_key:
            for fn, pn in inner.items():
                if _norm_simple(fn) == v_key:
                    return pn
            for fn, pn in inner.items():
                f_norm = _norm_simple(fn)
                if v_key.startswith(f_norm) or f_norm.startswith(v_key) or (f_norm in v_key) or (v_key in f_norm):
                    return pn
            if " " in v_key:
                first = v_key.split(" ", 1)[0]
                for fn, pn in inner.items():
                    if _norm_simple(fn).startswith(first):
                        return pn
            hit = "Unbekannt"
            break

    return hit or "Unbekannt"

# -------------------------------
# Styling
# -------------------------------
def apply_styles(sheet):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    name_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    data_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_fill_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    total_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    medium_border = Border(
        left=Side(style='medium', color='1F4E78'),
        right=Side(style='medium', color='1F4E78'),
        top=Side(style='medium', color='1F4E78'),
        bottom=Side(style='medium', color='1F4E78')
    )

    is_first_in_block = True
    alternate_row = False

    for row_idx, row in enumerate(sheet.iter_rows(min_col=1, max_col=5), start=1):
        first_cell_value = str(row[0].value).strip() if row[0].value else ""

        if "Gesamtverdienst" in first_cell_value:
            for cell in row:
                cell.fill = total_fill
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = medium_border
                if cell.column == 5:
                    cell.number_format = '#,##0.00 €'
            is_first_in_block = True
            alternate_row = False

        elif is_first_in_block and first_cell_value:
            for cell in row:
                cell.fill = name_header_fill
                cell.font = Font(bold=True, size=13, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = medium_border
            is_first_in_block = False
            alternate_row = False

        elif first_cell_value and not any(char.isdigit() for char in first_cell_value[:10]):
            if any(keyword in first_cell_value for keyword in ["Datum", "Tour", "LKW", "Art", "Verdienst"]):
                for cell in row:
                    cell.fill = subheader_fill
                    cell.font = Font(bold=True, size=10, color="1F4E78")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                alternate_row = False
            else:
                for cell in row:
                    cell.fill = name_header_fill
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = thin_border

        else:
            fill_color = data_fill_white if alternate_row else data_fill_light
            for cell in row:
                cell.fill = fill_color
                cell.font = Font(size=10, color="2C3E50")
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = thin_border
                if cell.column == 5:
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = '#,##0.00 €'
                        if cell.value > 0:
                            cell.font = Font(size=10, color="70AD47", bold=True)
                    except (ValueError, TypeError):
                        pass

            alternate_row = not alternate_row
            if not first_cell_value:
                is_first_in_block = True
                alternate_row = False

    column_min_widths = {1: 35, 2: 18, 3: 15, 4: 15, 5: 18}
    for col_idx, col in enumerate(sheet.columns, start=1):
        max_length = max(len(str(cell.value) or "") for cell in col)
        col_letter = get_column_letter(col[0].column)
        calculated_width = max_length + 6
        min_width = column_min_widths.get(col_idx, 12)
        adjusted_width = max(calculated_width, min_width)
        adjusted_width = min(adjusted_width, 65)
        sheet.column_dimensions[col_letter].width = adjusted_width

    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 20

    sheet.row_dimensions[1].hidden = True
    sheet.freeze_panes = "A3"

def format_date_with_german_weekday(date: pd.Timestamp) -> str:
    wochentage_mapping = {
        "Monday": "Montag", "Tuesday": "Dienstag", "Wednesday": "Mittwoch",
        "Thursday": "Donnerstag", "Friday": "Freitag", "Saturday": "Samstag",
        "Sunday": "Sonntag"
    }
    english_weekday = date.strftime("%A")
    german_weekday = wochentage_mapping.get(english_weekday, english_weekday)
    original_kw = int(date.strftime("%W"))
    adjusted_kw = original_kw + 1 if original_kw < 53 else 1
    return date.strftime(f"%d.%m.%Y ({german_weekday}, KW{adjusted_kw})")

def add_summary(sheet, summary_data, start_col=9, month_name=""):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    cell_fill_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    total_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    medium_border = Border(
        left=Side(style='medium', color='1F4E78'),
        right=Side(style='medium', color='1F4E78'),
        top=Side(style='medium', color='1F4E78'),
        bottom=Side(style='medium', color='1F4E78')
    )

    c1 = sheet.cell(row=2, column=start_col, value="Auszahlung Monat:")
    c1.font = Font(bold=True, size=14, color="FFFFFF")
    c1.alignment = Alignment(horizontal="right", vertical="center")
    c1.fill = header_fill
    c1.border = medium_border

    c2 = sheet.cell(row=2, column=start_col + 1, value=month_name or "Unbekannt")
    c2.font = Font(bold=True, size=14, color="FFFFFF")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.fill = header_fill
    c2.border = medium_border

    c3 = sheet.cell(row=2, column=start_col + 2, value="")
    c3.font = Font(bold=True, size=14, color="FFFFFF")
    c3.alignment = Alignment(horizontal="left", vertical="center")
    c3.fill = header_fill
    c3.border = medium_border

    headers = ["Name", "Personalnummer", "Gesamtverdienst (€)"]
    for i, h in enumerate(headers, start=start_col):
        cell = sheet.cell(row=3, column=i, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = subheader_fill
        cell.border = medium_border

    summary_data.sort(key=lambda x: x[2], reverse=True)

    for r, (name, personalnummer, total) in enumerate(summary_data, start=4):
        fill_color = cell_fill_white if r % 2 == 0 else cell_fill_light

        nc = sheet.cell(row=r, column=start_col, value=name)
        nc.font = Font(bold=True, size=11, color="2C3E50")
        nc.alignment = Alignment(horizontal="left", vertical="center")
        nc.fill = fill_color
        nc.border = thin_border

        pn_cell = sheet.cell(row=r, column=start_col + 1)
        if personalnummer != "Unbekannt":
            pn_cell.value = int(personalnummer)
            pn_cell.number_format = '00000000'
        else:
            pn_cell.value = personalnummer
        pn_cell.font = Font(size=10, color="5A6C7D")
        pn_cell.alignment = Alignment(horizontal="center", vertical="center")
        pn_cell.fill = fill_color
        pn_cell.border = thin_border

        tc = sheet.cell(row=r, column=start_col + 2, value=float(total))
        tc.font = Font(bold=True, size=11, color="70AD47")
        tc.fill = fill_color
        tc.border = thin_border
        tc.number_format = '#,##0.00 €'
        tc.alignment = Alignment(horizontal="right", vertical="center")

    total_row = len(summary_data) + 4

    lab = sheet.cell(row=total_row, column=start_col, value="GESAMTSUMME")
    lab.font = Font(bold=True, size=12, color="FFFFFF")
    lab.alignment = Alignment(horizontal="right", vertical="center")
    lab.fill = total_fill
    lab.border = medium_border

    empty_cell = sheet.cell(row=total_row, column=start_col + 1, value="")
    empty_cell.fill = total_fill
    empty_cell.border = medium_border

    sumcell = sheet.cell(row=total_row, column=start_col + 2, value=sum(x[2] for x in summary_data))
    sumcell.number_format = '#,##0.00 €'
    sumcell.font = Font(bold=True, size=12, color="FFFFFF")
    sumcell.fill = total_fill
    sumcell.border = medium_border
    sumcell.alignment = Alignment(horizontal="right", vertical="center")

    for row in range(2, total_row + 1):
        sheet.row_dimensions[row].height = 22

    sheet.column_dimensions[get_column_letter(start_col)].width = 28
    sheet.column_dimensions[get_column_letter(start_col + 1)].width = 20
    sheet.column_dimensions[get_column_letter(start_col + 2)].width = 22

# -------------------------------
# App
# -------------------------------
def main():
    st.title("Zulage - Sonderfahrzeuge - Ab 2025")

    uploaded_files = st.file_uploader(
        "Lade eine oder mehrere Excel-Dateien hoch",
        type=["xlsx", "xls"],
        accept_multiple_files=True
    )

    if uploaded_files:
        all_data = pd.DataFrame()

        for uploaded_file in uploaded_files:
            try:
                df = pd.read_excel(uploaded_file, sheet_name="Touren", header=0)

                # AZ-Zeilen filtern (Spalte 14 -> Index 13)
                mask = df.iloc[:, 13].astype(str).str.contains(r'(?i)\bAZ\b', na=False)
                filtered_df = df[mask].copy()

                if not filtered_df.empty:
                    filtered_df["Datum"] = pd.to_datetime(filtered_df.iloc[:, 14], format="%d.%m.%Y", errors="coerce")
                    filtered_df = filtered_df[filtered_df["Datum"] >= pd.Timestamp("2025-01-01")]

                if filtered_df.empty:
                    st.warning(f"Keine passenden Daten in der Datei {uploaded_file.name} gefunden.")
                    continue

                # ------------------------------------------------------------
                # NAMENLOGIK (wie gewünscht):
                # - Wenn D und E voll -> D/E verwenden
                # - Wenn D und E leer -> G/H verwenden
                # - NICHT beide nehmen
                # ------------------------------------------------------------
                # 0=Tour, 3=D Nachname, 4=E Vorname, 6=G Nachname2, 7=H Vorname2,
                # 10=LKW1, 11=LKW, 12=Art, 14=Datum
                columns_to_extract = [0, 3, 4, 6, 7, 10, 11, 12, 14]
                tmp = filtered_df.iloc[:, columns_to_extract].copy()
                tmp.columns = ["Tour", "Nachname_DE", "Vorname_DE", "Nachname_GH", "Vorname_GH", "LKW1", "LKW", "Art", "Datum"]

                def _clean_cell(x) -> str:
                    if pd.isna(x):
                        return ""
                    return str(x).replace("\u00a0", " ").strip()

                for c in ["Nachname_DE", "Vorname_DE", "Nachname_GH", "Vorname_GH"]:
                    tmp[c] = tmp[c].apply(_clean_cell)

                rows = []
                for _, r in tmp.iterrows():
                    de_ok = (r["Nachname_DE"] != "" and r["Vorname_DE"] != "")
                    if de_ok:
                        nn, vn = r["Nachname_DE"], r["Vorname_DE"]
                    else:
                        # nur wenn D/E leer -> G/H
                        nn, vn = r["Nachname_GH"], r["Vorname_GH"]

                    # Wenn auch G/H leer ist -> Zeile überspringen
                    if (nn == "" or vn == ""):
                        continue

                    rows.append({
                        "Tour": r["Tour"],
                        "Nachname": nn,
                        "Vorname": vn,
                        "LKW1": r["LKW1"],
                        "LKW": r["LKW"],
                        "Art": r["Art"],
                        "Datum": r["Datum"],
                    })

                extracted = pd.DataFrame(rows)

                if extracted.empty:
                    st.warning(f"AZ gefunden, aber keine verwertbaren Namen (D/E oder G/H) in {uploaded_file.name}.")
                    continue

                # LKW normalisieren + Art bestimmen
                extracted["LKW"] = extracted["LKW"].apply(lambda x: f"E-{x}" if pd.notnull(x) else x)
                extracted["Art"] = extracted["LKW"].apply(
                    lambda x: define_art(int(str(x).split("-")[1]))
                    if pd.notnull(x) and "-" in str(x) and str(x).split("-")[1].isdigit()
                    else "Unbekannt"
                )

                extracted["Datum"] = pd.to_datetime(extracted["Datum"], format="%d.%m.%Y", errors="coerce")

                # Tour ggf. aus Spalte Q (Index 16)
                if "Tour" in extracted.columns and filtered_df.shape[1] > 16:
                    extracted["Tour"] = extracted["Tour"].fillna(filtered_df.iloc[:, 16])

                # Verdienst
                def calculate_earnings(row):
                    earnings = 0
                    candidates = []

                    if pd.notnull(row["LKW1"]) and str(row["LKW1"]).isdigit():
                        candidates.append(int(row["LKW1"]))

                    if pd.notnull(row["LKW"]) and "-" in str(row["LKW"]):
                        tail = str(row["LKW"]).split("-")[1]
                        if tail.isdigit():
                            candidates.append(int(tail))

                    for v in candidates:
                        if v in [602, 156]:
                            earnings += 40
                        elif v in [620, 350, 520, 266, 458, 548, 541, 542, 543, 558]:
                            earnings += 20

                    return earnings

                extracted["Verdienst"] = extracted.apply(calculate_earnings, axis=1)
                extracted["Monat"] = extracted["Datum"].dt.month
                extracted["Jahr"] = extracted["Datum"].dt.year

                all_data = pd.concat([all_data, extracted], ignore_index=True)

            except Exception as e:
                st.error(f"Fehler beim Einlesen der Datei {uploaded_file.name}: {e}")

        if not all_data.empty:
            output_file = "touren_auswertung_korrekt.xlsx"
            try:
                with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                    sorted_data = all_data.sort_values(by=["Jahr", "Monat", "Nachname", "Vorname"])

                    for year, month in sorted_data[["Jahr", "Monat"]].drop_duplicates().values:
                        month_data = sorted_data[
                            (sorted_data["Jahr"] == year) & (sorted_data["Monat"] == month)
                        ].copy()
                        if month_data.empty:
                            continue

                        sheet_name = f"{get_german_month_name(month)} {year}"
                        sheet_data = []
                        summary_data = []

                        for (nachname, vorname), group in month_data.groupby(["Nachname", "Vorname"], dropna=False):
                            vn = (vorname or "").strip()
                            nn = (nachname or "").strip()

                            total_earnings = float(group["Verdienst"].sum())
                            personalnummer = get_personalnummer(nn, vn)

                            summary_data.append([f"{vn} {nn}".strip(), personalnummer, total_earnings])

                            sheet_data.append([f"{vn} {nn}".strip(), "", "", "", ""])
                            sheet_data.append(["Datum", "Tour", "LKW", "Art", "Verdienst"])

                            for _, row in group.iterrows():
                                dt = pd.to_datetime(row["Datum"]) if pd.notnull(row["Datum"]) else pd.NaT
                                formatted_date = format_date_with_german_weekday(dt) if pd.notnull(dt) else ""
                                sheet_data.append([
                                    formatted_date,
                                    row["Tour"],
                                    row["LKW"],
                                    row["Art"],
                                    float(row["Verdienst"])
                                ])

                            sheet_data.append(["Gesamtverdienst", "", "", "", total_earnings])
                            sheet_data.append([])

                        pd.DataFrame(sheet_data).to_excel(writer, index=False, sheet_name=sheet_name[:31])
                        sheet = writer.sheets[sheet_name[:31]]

                        add_summary(sheet, summary_data, start_col=9, month_name=sheet_name)
                        apply_styles(sheet)

                with open(output_file, "rb") as fh:
                    st.download_button(
                        label="Download Auswertung",
                        data=fh,
                        file_name="Zulage_Sonderfahrzeuge_2025.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"Fehler beim Exportieren der Datei: {e}")
        else:
            st.info("Keine Daten gefunden (nach AZ-Filter & Datum >= 01.01.2025).")

if __name__ == "__main__":
    main()
